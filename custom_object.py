import csv 
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font,Border,Side,Alignment

class Adress():
    """Class used to represent an Adress"""
    def __init__(self,name:str,street:str,postal:str,city:str) -> None:
        self.name=name
        self.street = street
        self.postal=postal
        self.city=city

    def getFullString(self) -> str:
        return(self.name+' '+self.street+' '+self.postal+' '+self.city)

    def getStreetString(self) -> str:
        return(self.street+' '+self.postal+' '+self.city)

class Travel(): # A modifier
    """A class used to represent a Travel"""
    def __init__(self,date:str,start_adress:Adress,end_adress:Adress,distance:str,price:str,rtrn_state:str) -> None:
        self.date = date
        self.start_adress = start_adress
        self.end_adress = end_adress
        self.distance = distance
        self.price = price
        self.rtrn_state = rtrn_state
    
    def getRow(self) -> list:
        return([self.date,self.start_adress.name,self.start_adress.street,self.start_adress.postal,self.start_adress.city,
                self.end_adress.name,self.end_adress.street,self.end_adress.postal,self.end_adress.city,
                self.distance,self.price,self.rtrn_state])
    
class Data():
    """A class used to get and set saved file datas"""    
    def __init__(self) -> None:
        self.historic_root = './data/historic.csv'
        self.adress_root = './data/adress.csv'
        self.user_data_root = './data/user_data.csv'
        self.travel_root = './data/travel.csv'
        self.ndf_excel_root = './data/NDF.xlsx'
    
    def getDataList(self,root) -> list[list[str]]:
        # return a list of file row with header
        with open(root,'r')as csv_file:
            reader = csv.reader(csv_file,delimiter=';')
            list_reader = list(reader)
        return(list_reader) 
    
    def getTravelList(self,root) -> list[Travel]:
        with open(root,'r')as csv_file:
            reader = csv.DictReader(csv_file,delimiter=';')
            list_reader = list(reader)
            travel_list = []
            for row in list_reader:
                date = row['date']
                start_adress = Adress(row['start_name'],row['start_street'],row['start_postal'],row['start_city'])
                end_adress = Adress(row['end_name'],row['end_street'],row['end_postal'],row['end_city'])
                distance = row['distance']
                price = row['price']
                rtrn_state = row['rtrn_state']
                travel = Travel(date,start_adress,end_adress,distance,price,rtrn_state)
                travel_list.append(travel)
        return(travel_list) 

    def getAdressList(self,root) -> list[Adress]:
        with open(root,'r')as csv_file:
            adress_list = []
            reader = csv.DictReader(csv_file,delimiter=';')
            list_reader = list(reader)
            for row in list_reader:
                name = row['name']
                street = row['street']
                postal = row['postal']
                city = row['city']
                adress = Adress(name,street,postal,city)
                adress_list.append(adress)
        return(adress_list)   
    
    def writeData(self,root:str,data:list) -> None:
        with open(root,'w',newline="")as csv_file:
            writer = csv.writer(csv_file,delimiter=';')
            for row in data:
                writer.writerow(row)

    def saveAdress(self,adress:Adress) -> None:
        """Save adress in adress.csv if it's new adress"""
        data = self.getDataList(self.adress_root)
        row = [adress.name,adress.street,adress.postal,adress.city]
        if not(row in data):
            data.append(row)
            self.writeData(self.adress_root,data)
            print(f'save new adress : {row} in {self.adress_root}')

    def deleteTravel(self,travel:Travel) -> None:
        row_list = self.getDataList(self.historic_root)
        row = travel.getRow()
        if row in row_list:
            index = row_list.index(row)
            row_list.pop(index)
            print(f'delete old row at index {index} in {self.historic_root}')
            self.writeData(self.historic_root,row_list)

    def saveToHistoric(self,old_travel:Travel,new_travel:Travel) -> None:
        """Sauvegarde un travel dans historic.csv"""
        row_list= self.getDataList(self.historic_root)
        new_row = new_travel.getRow()
        
        if old_travel != None:
            old_row = old_travel.getRow()
            # Si on modifie l'élément on commence par supprimer l'ancien élément
            if old_row in row_list:
                index = row_list.index(old_row)
                row_list.pop(index)
                print(f'delete old row at index {index} in {self.historic_root}')

        # On parcour la liste des travel a l'envers, pour chaque date on compare
        if len(row_list) > 1:
            for row in reversed(row_list):
                index = row_list.index(row)
                # Si on a parcouru toute la liste sans rien comparer
                if index == 0:
                    row_list.insert(1,new_row)
                    break
                else:
                    date = datetime.strptime(row[0],'%d/%m/%Y')
                    new_date = datetime.strptime(new_row[0],'%d/%m/%Y')
                    if new_date >= date:
                        row_list.insert(index+1,new_row)
                        break
        # Démarage
        else:
            row_list.append(new_row)

        self.writeData(self.historic_root,row_list)
        print('save new travel in historic')

    def saveToTravel(self,travel:Travel) -> None:
        data = self.getDataList(self.travel_root)
        start_name = travel.start_adress.name
        end_name = travel.end_adress.name
        distance = travel.distance
        for row in data[1:]:
            if not(start_name in row and end_name in row): # Si les deux éléments ne sont pas dans la liste
                new_row = [start_name,end_name,distance]
                data.append(new_row)
                self.writeData(self.travel_root,data)
                print(f'save new travel {new_row} in {self.travel_root}')
                break

    def getTravelDistance(self,start_adress_name:str,end_adress_name:str) -> str|bool:
        """Check if travel is in travel.csv and return distance"""
        travel_list = self.getDataList(self.travel_root)[1:]
        for row in travel_list:
            if start_adress_name in row and end_adress_name in row:
                return row[2]
        return False

class Roots():
    def __init__(self) -> None:
        self.historic = './data/historic.csv'
        self.adress = './data/adress.csv'
        self.user_data = './data/user_data.csv'
        self.travel = './data/travel.csv'
        self.ndf_excel = './data/NDF.xlsx'

class Excel():
    """Class used to generate excel file"""
    def __init__(self,start_date,end_date) -> None: 
        self.data = Data()
        self.start_date = start_date
        self.end_date = end_date
        self.root = Roots()

        self.header_row = 6
        self.start_tab_row = self.header_row+1

        self.tab_dict = self.getTabDict(self.start_tab_row,self.start_date,self.end_date)
        
        self.end_tab_row = self.start_tab_row+len(list(self.tab_dict.values())[0])

        self.wb = Workbook()
        self.sh = self.wb.active
        
        side = Side(border_style="thin", color="000000")
        self.border = Border(top=side, left=side, right=side, bottom=side)
        self.bold_font = Font(bold=True)
        
        self.setColumnDim()
        self.setSheet(self.start_date,self.end_date,self.border,self.bold_font)
        self.setTab(self.tab_dict,self.border,self.bold_font)
        self.setBottomTab(self.end_tab_row,self.border,self.bold_font)
        
    def getTabDict(self,start_row,start_date,end_date) -> dict:
        start_date = datetime.strptime(start_date,'%d/%m/%Y')
        end_date = datetime.strptime(end_date,'%d/%m/%Y')

        # Lecture du fichier
        df = pd.read_csv(self.root.historic,sep=';')
        df_columns = ['date','end_name','end_street','end_postal','end_city','distance','price']
        nb_rows = df.shape[0]

        # Dictionnaire qui contient les données à afficher
        tab_dict={'Date':[],
              'Ecole + Trajet':[],
              'Adresse':[],
              'CP':[],
              'Localité':[],
              'Km/AR':[],
              '€/km':[]}

        # Pour chaque ligne du fichier csv : 
            # si la date appartient à l'intervale -> ajoute les données de la ligne au dictionnaire
        for row_index in range(nb_rows):
            date = datetime.strptime(df['date'][row_index],'%d/%m/%Y')
            if date >= start_date and date <= end_date:
                for csv_key,dict_key in zip (df_columns,tab_dict.keys()):
                    if dict_key == 'Km/AR':
                        if df['rtrn_state'][row_index]:
                            distance = float(df['distance'][row_index])*2
                            tab_dict[dict_key].append(str(distance))
                        else: tab_dict[dict_key].append(df[csv_key][row_index])
                    else: tab_dict[dict_key].append(df[csv_key][row_index])
            else: pass

        # Ajoute la colonne Total au tableau
        tab_dict['Total'] = []
        for i in range(len(list(tab_dict.values())[0])):
            row_index = start_row+1+i 
            tab_dict['Total'].append(f'=F{row_index}*G{row_index}')
        return(tab_dict)

    def setSheet(self,start_date,end_date,border:Border,font:Font) -> None:
        df = pd.read_csv(self.root.user_data,sep=';',index_col=False)
        user_name = df['user_name'][0]
        code = df['code'][0]
        user_adress = df['user_adress'][0]
        user_bank = df['user_bank'][0]

        title = f'Frais de déplacement du {start_date} au {end_date}'

        head = {'A1':title,
                
                'G1':code,
                'B3':user_name,
                'C3':user_adress,
                'B4':'Compte bancaire',
                'C4':user_bank}
        
        for key in head.keys():
            self.sh[key] = head[key]
            self.sh[key].border = border
            self.sh[key].font = font

        self.sh["A1"].alignment = Alignment(horizontal="center")

        self.sh.merge_cells('A1:F1')
        self.sh.merge_cells('G1:H1')
        self.sh.merge_cells('C3:H3')
        self.sh.merge_cells('C4:H4')
    
    def setColumnDim(self) -> None:
        for dim,column in zip ([15,40,30,10,30,10,10,10],['A','B','C','D','E','F','G','H']):
            self.sh.column_dimensions[column].width = dim

    def setTab(self,tab_dict:dict,border:Border,font:Font) -> None:
        # Ecrit le tableau de donné
        for key,column in zip (self.tab_dict.keys(),range(1,len(tab_dict)+1)): # Ecrit le header
            cell = self.sh.cell(row = self.start_tab_row,column = column, value = key)
            cell.border = border
            cell.font = font
            for row in range(len(list(tab_dict.values())[0])): # Ecrit les données
                cell = self.sh.cell(row = self.start_tab_row+1+row, column = column, value = tab_dict[key][row])
                cell.border = border

    def setBottomTab(self,start_bottom_row,border:Border,font:Font) -> None:
        tab = {'Déplacement divers':start_bottom_row+1,
               'Total parkin':start_bottom_row+2,
               'Total Déplacement':start_bottom_row+3}
        
        for key in tab.keys():
            self.sh[f'A{tab[key]}'] = key
            self.sh[f'A{tab[key]}'].border = border
            self.sh[f'H{tab[key]}'].border = border
            self.sh[f'A{tab[key]}'].font = font
            self.sh.merge_cells(f'A{tab[key]}:G{tab[key]}')
        
        total = f'=sum(H{self.start_tab_row+1}:H{self.end_tab_row})'
        self.sh.cell(row=tab['Total Déplacement'], column = len(self.tab_dict),value = total)

    def save(self) -> None:
        self.wb.save(self.data.ndf_excel_root)

